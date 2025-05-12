from flask import Flask, render_template, request, redirect, send_file, session, url_for, flash, session, abort
import os
from config import Config
import uuid
from modules.process import PDFEstatementProcessor
import pandas as pd
from extensions import db
from sqlalchemy import or_, asc
from datetime import datetime
from io import StringIO, BytesIO
import traceback
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension
import ast
import re

app = Flask(__name__)

app.config.from_object(Config)
db.init_app(app)

from models import *

app.secret_key = 'your-secret-key'
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(BASE_DIR, 'output')

processor = PDFEstatementProcessor()

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

@app.before_request
def global_middleware():
    if request.endpoint in ('login', 'static'):
        return
    
    if request.path == '/login':
        if session.get('user_id'):
            return redirect('/dashboard')

    if request.path == '/':
        if session.get('user_id'):
            return redirect('/dashboard')
        else:
            return redirect('/login')

    protected_routes = {
        '/user': ['super_admin']
    }

    allowed_roles = protected_routes.get(request.path)
    if allowed_roles:
        user_role = session.get('role')
        if not user_role or user_role not in allowed_roles:
            return abort(403)
        
@app.route('/converter', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        session.pop('data', None)
        session.pop('data_month', None)
        session.pop('output_filename', None)
        files = request.files.getlist('pdf_file[]')
        niks = request.form.get('nik')
        mobile_phones = request.form.get('mobile_phone')
        names = request.form.get('name')
        sources = request.form.getlist('source[]') 
        emails = request.form.get('email')
        emails_pass = request.form.get('email_pass')
        users = request.form.get('user')
        users_pass = request.form.get('user_pass')
        sources_per_batch = request.form.getlist('source_per_batch[]')
        user_list = request.files.get('user_list')

        if not files or not niks or not mobile_phones or not names or not sources or not user_list:
            return "All fields are required."
        
        mobile_phones, niks = f"{''.join(re.findall(r'\d+', mobile_phones))}", f"{''.join(re.findall(r'\d+', niks))}"
        
        file_list = []
        for file in files:
            file_list.append(file.filename)

        final_sources = [
            source
            for source, count in zip(sources, sources_per_batch)
            for _ in range(int(count))
        ]

        account_list = pd.read_excel(user_list, engine='openpyxl')
                
        combined_results = []

        for idx, file in enumerate(files):
            if file and file.filename.endswith('.pdf'):
                filename = f"{uuid.uuid4()}.pdf"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                try:
                    data = TransferUser.query.all()
                    bank_code = BankCode.query.filter(BankCode.bank_name.like(f"%{final_sources[idx]}%")).order_by(asc(BankCode.id)).first()
                    is_av, result, _ = processor.process_pdf_file(filepath, BankCode, data, names, niks, mobile_phones, final_sources[idx], bank_code.bank_code, emails, emails_pass, users, users_pass, account_list)
                    if is_av:
                        combined_results.append(result)
                except Exception as e:
                    error_details = traceback.format_exc()
                    print(error_details)
                    return f"Error processing CSV data: {str(e)}"
            else:
                return f"Invalid file format for {file.filename}. Only PDFs are supported."

        if combined_results:
            final_df = pd.concat(combined_results, ignore_index=True)
            final_df = final_df[final_df.notna().all(axis=1)]
            final_df['% date'] = pd.to_datetime(final_df['% date'], format='%d-%b-%Y', errors='coerce')
            final_df = final_df.sort_values(by='% date')
            final_df['% date'] = pd.to_datetime(
                final_df['% date'],
                format='%d/%m/%Y',
                errors='coerce'
            ).dt.strftime('%d-%b-%Y')
            table_html = final_df.to_html(classes='table table-striped', index=False)
            csv_data = final_df.to_csv(index=False)
            return render_template('converter/preview.html', title="Converter", name=names, nik=niks, mobile_phone=mobile_phones, data=csv_data, table=table_html, file_list=file_list, filename=f"export.xlsx", current_url=request.url)

        return "No valid files processed."
    
    banks = BankCode.query.all()
    return render_template('converter/index.html', current_url=request.url, title="Converter", banks=banks)


@app.get('/login')
def login_view():
    return render_template('auth/login.html')

@app.post('/login')
def login_process():
    email, password = request.form['email'], request.form['password']
    
    user = User.query.filter_by(email=email).first()
    if(user and user.verify_password(password)):
        session['user_id'] = user.id
        session['username'] = user.username
        session['email'] = user.email
        session['role'] = user.role
        return redirect('/dashboard')
    else:
        flash("The credentials doesn't match our records!", "message")
        return redirect(request.referrer or '/')

@app.post('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('username', None)
    session.pop('email', None)
    session.pop('role', None)
    flash("You’ve been logged out.", "message")
    return redirect('/')
    
@app.get('/dashboard')
def render_dashboard():
    return  render_template('dashboard/index.html', current_url=request.url, title="Dashboard")

@app.get('/user')
def render_user():
    users = User.query.filter(User.id != session.get('user_id')).all()
    return render_template('user/index.html', users=users, current_url=request.url, title="User Management")

@app.post('/user/<int:id>')
def update_user(id):
    method = request.form.get('_method', '').upper()
    if method == "PUT":
        username = request.form['username']
        email = request.form['email']
        password = request.form.get('password')
        role = request.form['role']

        user = User.query.get_or_404(id)
        user.username = username
        user.email = email
        user.role = role
        if password:
            user.password = password
        db.session.commit()

        flash("User updated!", "message")
        flash("true", "success")
        return redirect(request.referrer or '/dashboard')
    elif method == "DELETE":
        user = User.query.get_or_404(id)
        db.session.delete(user)
        db.session.commit()
        flash("User deleted!", "message")
        flash("true", "success")
        return redirect(request.referrer or '/dashboard')
    flash("Method not allowed!", "message")
    flash("false", "success")
    return redirect(request.referrer or '/dashboard')

@app.post('/user')
def create_user():
    username, email, password = request.form['username'], request.form['email'], request.form['password']

    if(not username or not email or not password):
       flash("Fill all of those credentials!", "message")
       flash("false", "success")
       return redirect(request.referrer or '/')
    
    prev_user = User.query.filter(or_(User.email == email, User.username == username)).first()
    if(prev_user):
        flash("The username or email already taken!", "message")
        flash("false", "success")
        return redirect(request.referrer or '/')
    
    user = User(
        email=email,
        username=username
    )
    user.password = password
    db.session.add(user)
    db.session.commit()

    flash("User created!", "message")
    flash("true", "success")
    return redirect(request.referrer or '/')


# @app.post('/preview')
# def preview():
#     if 'data' in session:
#         df = pd.read_json(session['data'], orient='split')
#         table_html = df.to_html(classes='table table-striped', index=False)
#         return render_template('converter/preview.html', title="Converter", table=table_html, month=session['data_month'], current_url=request.url, filename=session['output_filename'])
#     return redirect('/converter')


@app.post('/download/<filename>')
def download_file(filename):
    data = request.form.get('data')
    file_list = ast.literal_eval(request.form.get('file_list'))
    name = request.form.get('name')
    nik = request.form.get('nik')
    mobile_phone = request.form.get('mobile_phone')

    if data:
        try:
            csv_buffer = StringIO(data)

            df = pd.read_csv(csv_buffer, dtype={'% A Bank Code': str, '% B Bank Code': str, '% A NIK': str, '% B NIK': str, '% A Mobile': str, '% B Mobile': str, '% A number': str, '% B number': str})
            df['% nominal'] = df['% nominal'].str.replace(',', '', regex=False).astype(float)
            df['% date'] = df['% date'].apply(lambda x: datetime.strptime(x, '%d-%b-%Y'))

            output_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)

            workbook = xlsxwriter.Workbook(output_path)
            worksheet = workbook.add_worksheet('Sheet1')

            date_format = workbook.add_format({'num_format': 'dd-mmm-yy'})
            accounting_format = workbook.add_format({'num_format': '#,##0'})
            text_format = workbook.add_format({'num_format': '@'})
            header_format = workbook.add_format({
                'bold': True,
                'border': 1, 
                'bg_color': '#D9D9D9'
            })

            for col_idx, col_name in enumerate(df.columns):
                worksheet.write(0, col_idx, col_name, header_format)

            for row_idx, row in enumerate(df.itertuples(index=False), start=1):
                for col_idx, value in enumerate(row):
                    col_name = df.columns[col_idx]
                    if col_name == '% date':
                        worksheet.write_datetime(row_idx, col_idx, value, date_format)
                    elif col_name == '% nominal':
                        worksheet.write_number(row_idx, col_idx, value, accounting_format)
                    elif df.dtypes[col_name] == object:
                        worksheet.write_string(row_idx, col_idx, str(value), text_format)
                    else:
                        worksheet.write(row_idx, col_idx, value)

            worksheet.set_column('A:A', 15, date_format)
            worksheet.set_column('B:B', 10, accounting_format)
            for col in range(2, len(df.columns)):
                worksheet.set_column(col, col, 15, text_format)

            workbook.close()

            log = ConverterLog(
                user_id = session['user_id'],
                files = file_list,
                name = name, 
                nik = nik,
                mobile_phone = mobile_phone
            )
            db.session.add(log)
            db.session.commit()

            return send_file(output_path, as_attachment=True)

        except Exception as e:
            return f"Error processing CSV data: {str(e)}"

    return "No file available to download."

@app.get('/converter-log')
def converter_log():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    logs = ConverterLog.query.filter_by(user_id=session['user_id']).order_by(ConverterLog.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template('converter_log/index.html', title="Converter Log", logs=logs, current_url=request.url)


@app.get('/daftar-transfer/template')
def daftar_transfer_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank User Template"

    headers = [
        "% B number",
        "% B Bank Code",
        "% B Bank Name",
        "% B NIK",
        "% B name",
        "% B Mobile",
        "% B Email Pass",
        "% B User Pass"
    ]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
        cell.number_format = '@' 

        for row in range(2, 102):
            data_cell = ws.cell(row=row, column=col_num)
            data_cell.number_format = '@'

    widths = [15, 15, 20, 20, 25, 15, 25, 20]
    for i, width in enumerate(widths, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="template_bank_user.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
if __name__ == '__main__':
    app.run(debug=True)
